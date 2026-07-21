import re
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, LineChart, Reference

HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")  # Tailwind Slate-900
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")

# Column-name hints used to guess which numeric column represents money, so
# we know what to currency-format and what to chart — this is a heuristic,
# not something the user configures for a quick ad-hoc chat export.
MONEY_NAME_HINTS = ("amount", "revenue", "price", "spend", "cost", "total", "value", "sales", "profit")


class ExcelExporter:
    @staticmethod
    def _pick_money_column(df: pd.DataFrame, numeric_cols: list) -> str | None:
        for col in numeric_cols:
            if any(hint in col.lower() for hint in MONEY_NAME_HINTS):
                return col
        return numeric_cols[0] if numeric_cols else None

    @staticmethod
    def _write_styled_table(ws, df: pd.DataFrame, money_col: str | None = None):
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN

            column_letter = get_column_letter(col_num)
            try:
                max_length = max(df[column_title].astype(str).map(len).max(), len(str(column_title)))
            except ValueError:
                max_length = len(str(column_title))
            ws.column_dimensions[column_letter].width = min(max_length + 3, 40)

            if money_col and column_title == money_col:
                for row_num in range(2, len(df) + 2):
                    ws.cell(row=row_num, column=col_num).number_format = '"$"#,##0.00'

        ws.freeze_panes = "A2"
        if len(df) > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

            if money_col and money_col in df.columns:
                money_col_idx = list(df.columns).index(money_col) + 1
                col_letter = get_column_letter(money_col_idx)
                color_rule = ColorScaleRule(
                    start_type="min", start_color="FFC7CE",
                    mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                    end_type="max", end_color="C6EFCE",
                )
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(df) + 1}", color_rule)

    @staticmethod
    def export_to_excel(df: pd.DataFrame, sheet_name: str = "Data") -> BytesIO:
        """
        Generates an investor-ready Excel workbook: styled raw data (with
        currency formatting, conditional color-scale, autofilter, frozen
        header), plus — when the data shape supports it — a Summary sheet
        with a real, native, editable Excel chart (not a pasted image).
        """
        numeric_cols = list(df.select_dtypes(include="number").columns)
        categorical_cols = [
            c for c in df.select_dtypes(exclude="number").columns
            if df[c].nunique() <= 50
        ]
        money_col = ExcelExporter._pick_money_column(df, numeric_cols)

        wb = Workbook()
        data_ws = wb.active
        data_ws.title = sheet_name
        ExcelExporter._write_styled_table(data_ws, df, money_col=money_col)

        # --- Summary sheet with a native chart, if the shape supports it ---
        if money_col and categorical_cols:
            group_col = categorical_cols[0]
            summary_df = (
                df.groupby(group_col)[money_col]
                .agg(["sum", "mean", "count"])
                .reset_index()
                .rename(columns={"sum": f"Total {money_col}", "mean": f"Avg {money_col}", "count": "Row Count"})
            )
            summary_df = summary_df.sort_values(by=f"Total {money_col}", ascending=False)

            summary_ws = wb.create_sheet("Summary")
            ExcelExporter._write_styled_table(summary_ws, summary_df)

            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = f"Total {money_col} by {group_col}"
            chart.y_axis.title = money_col
            chart.x_axis.title = group_col

            data_ref = Reference(summary_ws, min_col=2, min_row=1, max_row=len(summary_df) + 1)
            cats_ref = Reference(summary_ws, min_col=1, min_row=2, max_row=len(summary_df) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.width = 18
            chart.height = 10
            summary_ws.add_chart(chart, f"A{len(summary_df) + 4}")

        # --- Trend sheet with a native line chart, if there's a date-like column ---
        date_col = None
        for c in df.columns:
            if c not in numeric_cols:
                try:
                    parsed = pd.to_datetime(df[c], errors="coerce", format="mixed")
                    if parsed.notna().mean() > 0.8:
                        date_col = c
                        break
                except Exception:
                    continue

        if money_col and date_col:
            trend_df = df.copy()
            trend_df["_period"] = pd.to_datetime(trend_df[date_col], errors="coerce", format="mixed").dt.to_period("M").astype(str)
            trend_summary = trend_df.groupby("_period")[money_col].sum().reset_index().rename(
                columns={"_period": "Period", money_col: f"Total {money_col}"}
            )

            trend_ws = wb.create_sheet("Trend")
            ExcelExporter._write_styled_table(trend_ws, trend_summary)

            line_chart = LineChart()
            line_chart.title = f"{money_col} Over Time"
            line_chart.y_axis.title = money_col
            line_chart.x_axis.title = "Period"
            data_ref = Reference(trend_ws, min_col=2, min_row=1, max_row=len(trend_summary) + 1)
            cats_ref = Reference(trend_ws, min_col=1, min_row=2, max_row=len(trend_summary) + 1)
            line_chart.add_data(data_ref, titles_from_data=True)
            line_chart.set_categories(cats_ref)
            line_chart.width = 18
            line_chart.height = 10
            trend_ws.add_chart(line_chart, f"A{len(trend_summary) + 4}")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
